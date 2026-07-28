"""
Génération du fichier Excel de recommandations de cadence — livrable
téléchargeable pour le responsable production.

N'ajoute aucune logique métier : consomme directement les objets déjà
produits par src.recommend.recommender.recommend_cadences (Phase 7).
Ce module ne fait que la mise en forme du classeur.
"""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_COLOR_HEADER_BG = "3A8F17"
_COLOR_HEADER_TEXT = "FFFFFF"
_COLOR_FIABLE_BG = "E7FDDF"
_COLOR_FIABLE_TEXT = "40802F"
_COLOR_AVERIFIER_BG = "FDF3D6"
_COLOR_AVERIFIER_TEXT = "8A6D1A"

_HEADERS = [
    "Produit",
    "Machine",
    "Fiable",
    "Cadence théorique actuelle (pcs/min)",
    "Cadence recommandée (pcs/min)",
    "Écart vs théorique (%)",
    "TRS moyen de référence (%)",
    "OF utilisés / disponibles",
    "Justification",
]


def _autofit_columns(ws: Worksheet, min_width: int = 10, max_width: int = 60) -> None:
    for col_idx, _ in enumerate(_HEADERS, start=1):
        letter = get_column_letter(col_idx)
        longest = max(
            (len(str(cell.value)) for cell in ws[letter] if cell.value is not None),
            default=min_width,
        )
        ws.column_dimensions[letter].width = max(min_width, min(longest + 2, max_width))


def build_recommendations_workbook(recommendations: list) -> bytes:
    """
    Construit le classeur Excel des recommandations et retourne son
    contenu binaire, prêt à être streamé en réponse HTTP.

    `recommendations` : liste d'objets CadenceRecommendation
    (src.recommend.models.CadenceRecommendation), fiables ou non.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Recommandations cadences"

    ws.append(_HEADERS)
    header_font = Font(name="Arial", bold=True, color=_COLOR_HEADER_TEXT, size=11)
    header_fill = PatternFill(start_color=_COLOR_HEADER_BG, end_color=_COLOR_HEADER_BG, fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    body_font = Font(name="Arial", size=10)
    for rec in recommendations:
        row = [
            rec.produit,
            rec.machine,
            "Fiable" if rec.fiable else "À vérifier",
            round(rec.cadence_theorique_actuelle, 2) if rec.cadence_theorique_actuelle is not None else "N/A",
            round(rec.cadence_recommandee, 2) if rec.cadence_recommandee is not None else "N/A",
            rec.ecart_vs_theorique_pct if rec.ecart_vs_theorique_pct is not None else "N/A",
            rec.trs_moyen_reference if rec.trs_moyen_reference is not None else "N/A",
            f"{rec.n_of_utilises} / {rec.n_of_disponibles}",
            " ".join(rec.justification),
        ]
        ws.append(row)

        current_row = ws.max_row
        fiable_cell = ws.cell(row=current_row, column=3)
        if rec.fiable:
            fiable_cell.fill = PatternFill(start_color=_COLOR_FIABLE_BG, end_color=_COLOR_FIABLE_BG, fill_type="solid")
            fiable_cell.font = Font(name="Arial", bold=True, color=_COLOR_FIABLE_TEXT, size=10)
        else:
            fiable_cell.fill = PatternFill(
                start_color=_COLOR_AVERIFIER_BG, end_color=_COLOR_AVERIFIER_BG, fill_type="solid"
            )
            fiable_cell.font = Font(name="Arial", bold=True, color=_COLOR_AVERIFIER_TEXT, size=10)

        for col_idx in range(1, len(_HEADERS) + 1):
            if col_idx == 3:
                continue
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = body_font
            if col_idx == len(_HEADERS):
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}{ws.max_row}"
    _autofit_columns(ws)

    info_ws = wb.create_sheet("À propos")
    info_ws["A1"] = "Rapport de recommandation de cadence — Kadansa"
    info_ws["A1"].font = Font(name="Arial", bold=True, size=13)
    info_ws["A2"] = f"Généré le {datetime.now().strftime('%Y-%m-%d à %H:%M')}"
    info_ws["A2"].font = Font(name="Arial", italic=True, size=10)
    info_ws["A4"] = (
        "Les recommandations marquées « À vérifier » reposent sur un historique "
        "encore insuffisant (moins de 3 OF exploitables) et ne doivent pas être "
        "appliquées telles quelles sans validation terrain."
    )
    info_ws["A4"].alignment = Alignment(wrap_text=True)
    info_ws.column_dimensions["A"].width = 90
    info_ws.row_dimensions[4].height = 45

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()